"""
漏洞扫描 AI 辅助修复 — 解析服务（第一阶段：只生成待审批任务，不下发执行）

流程（parse_import 后台任务）：
  1. openpyxl 读 xlsx（列：IP, DNS_Calculation, QID, Title, Results, Solution）
  2. 逐行落 vuln_findings + 资产匹配（hostname 精确 > ip 精确 > 短名 fuzzy > unmatched）
  3. 生成 remediation_tasks：
     a. remediation_rules 有该 QID（active/draft）→ 直接套用规则
     b. 无规则 → 调 openclaw 网关 LLM 解析 Title+Solution → 结构化 JSON，
        同时写一条 status=draft 的规则草稿供人工转正
     c. 规则 disabled / LLM 不可用 → manual_review + high
  4. 风险分级策略见 classify_risk()
"""
import asyncio
import json
import logging
import re
from datetime import datetime
from pathlib import Path
from typing import Optional

import httpx
from openpyxl import load_workbook
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from app.core.config import get_settings
from app.core.database import AsyncSessionLocal
from app.models.models import Client
from app.models.vuln import (
    VulnScanImport, VulnFinding, RemediationTask, RemediationRule,
    FIX_TYPES,
)

logger = logging.getLogger(__name__)
settings = get_settings()

# 期望的表头（大小写不敏感，允许乱序/多余列）
EXPECTED_HEADERS = ["ip", "dns_calculation", "qid", "title", "results", "solution"]

# 安全软件关键词（EDR/杀软相关的升级/卸载一律 high）
SECURITY_SOFTWARE_KEYWORDS = [
    "symantec", "forticlient", "fortinet", "crowdstrike", "mcafee",
    "kaspersky", "trend micro", "trendmicro", "sentinelone", "sentinel one",
    "carbon black", "carbonblack", "sophos", "windows defender", "cortex xdr",
    "bitdefender", "eset", "qihoo", "360safe", "火绒", "huorong",
]

# EOL / 停止支持类关键词 → manual_review
EOL_KEYWORDS = [
    "eol", "end of life", "end-of-life", "end of support", "end-of-support",
    "obsolete", "no longer supported", "unsupported version",
]


# ── 风险分级 ──────────────────────────────────────────────────────────────────
def is_security_software(text: str) -> bool:
    t = (text or "").lower()
    return any(k in t for k in SECURITY_SOFTWARE_KEYWORDS)


def is_eol_vuln(title: str, solution: str = "") -> bool:
    t = f"{title} {solution}".lower()
    return any(k in t for k in EOL_KEYWORDS)


def classify_risk(fix_type: str, title: str = "", solution: str = "",
                  llm_confidence: Optional[str] = None) -> str:
    """
    默认风险分级策略：
      registry_fix（非安全软件）        = low
      patch_install / software_upgrade  = medium
      涉及 EDR/杀软 的升级/卸载          = high
      manual_review（EOL 类）/unsupported= high
      LLM confidence=low                → 一律升级为 high
    """
    text = f"{title} {solution}"
    sec = is_security_software(text)

    if fix_type in ("manual_review", "unsupported"):
        risk = "high"
    elif fix_type in ("software_upgrade", "software_uninstall") and sec:
        risk = "high"
    elif fix_type == "registry_fix":
        risk = "high" if sec else "low"
    elif fix_type in ("patch_install", "software_upgrade"):
        risk = "medium"
    elif fix_type == "software_uninstall":
        risk = "medium"
    else:
        risk = "medium"

    if llm_confidence == "low":
        risk = "high"

    # patch_install 最低风险等级可配置（默认 medium，保守环境可设 high）
    if fix_type == "patch_install" and risk not in ("high",):
        try:
            from app.core.config import get_settings
            min_risk = get_settings().VULN_PATCH_MIN_RISK
        except Exception:
            min_risk = "medium"
        _rank = {"low": 0, "medium": 1, "high": 2}
        if _rank.get(min_risk, 1) > _rank.get(risk, 0):
            risk = min_risk
    return risk


def action_summary(fix_type: str, action: Optional[dict]) -> str:
    """action_json 的一句话摘要（列表页展示用）"""
    a = action or {}
    try:
        if fix_type == "registry_fix":
            changes = a.get("changes") or []
            if changes:
                n = len(changes)
                first = changes[0]
                sample = f"{first.get('path', '?')}\\{first.get('value', '?')}"
                return f"注册表：设置 {n} 个键值（如 {sample}…）"
            # 兼容旧单键格式
            return f"注册表：{a.get('registry_path', '?')} → {a.get('value_name', '?')}={a.get('value_data', '?')}"
        if fix_type == "software_upgrade":
            tv = a.get("target_version") or "最新版"
            return f"升级 {a.get('software', '?')} 至 {tv}"
        if fix_type == "software_uninstall":
            return f"卸载 {a.get('software', '?')}"
        if fix_type == "patch_install":
            kbs = a.get("kb_ids") or []
            return f"安装补丁 {', '.join(kbs[:4])}{'…' if len(kbs) > 4 else ''}" if kbs else "安装系统补丁"
        if fix_type == "manual_review":
            return f"人工处理：{a.get('reason', a.get('description', ''))[:80]}"
        if fix_type == "unsupported":
            return f"暂不支持自动修复：{a.get('reason', '')[:80]}"
    except Exception:
        pass
    return (a.get("description") or "")[:80] if isinstance(a, dict) else ""


# ── 资产匹配 ──────────────────────────────────────────────────────────────────
async def load_asset_index(db: AsyncSession) -> dict:
    """一次性加载资产表索引（hostname 小写 / 短名 / ip）"""
    rows = (await db.execute(select(Client.id, Client.hostname, Client.ip))).all()
    by_hostname, by_short, by_ip = {}, {}, {}
    for cid, hostname, ip in rows:
        if hostname:
            h = hostname.strip().lower()
            by_hostname.setdefault(h, cid)
            by_short.setdefault(h.split(".")[0], cid)
        if ip:
            by_ip.setdefault(ip.strip(), cid)
    return {"hostname": by_hostname, "short": by_short, "ip": by_ip}


def match_asset(index: dict, dns_name: Optional[str], ip: Optional[str]):
    """返回 (asset_id, match_confidence)"""
    d = (dns_name or "").strip().lower()
    i = (ip or "").strip()
    if d and d in index["hostname"]:
        return index["hostname"][d], "exact_hostname"
    if i and i in index["ip"]:
        return index["ip"][i], "exact_ip"
    if d:
        short = d.split(".")[0]
        if short and short in index["short"]:
            return index["short"][short], "fuzzy"
    return None, "unmatched"


# ── LLM 解析（openclaw 网关，OpenAI 兼容 /v1）────────────────────────────────
LLM_SYSTEM_PROMPT = """你是企业漏洞修复专家。根据 Qualys 漏洞扫描条目的 Title 和 Solution，输出结构化修复建议。
只输出一个 JSON 对象，不要输出任何其它文字或 markdown 代码块。JSON 格式：
{
  "fix_type": "registry_fix|software_upgrade|software_uninstall|patch_install|manual_review|unsupported",
    "action": {
    // registry_fix: {"changes":[{"hive":"HKLM","path":"SYSTEM\\\\CurrentControlSet\\\\Services\\\\LanmanServer\\\\Parameters","value":"requiresecuritysignature","type":"REG_DWORD","data":1}], "requires_reboot": false, "description":"..."}
    //   ⚠ 重要：一个加固项若涉及多个注册表键（典型如 SMB Signing QID 90043，必须同时设置
    //      RequireSecuritySignature=1 与 EnableSecuritySignature=1），必须把每一个键都作为
    //      changes 数组里的一项逐一列出，严禁只输出其中一个键。
    // software_upgrade:  {"software": "...", "target_version": "...", "download_hint": "...", "description": "..."}
    // software_uninstall:{"software": "...", "description": "..."}
    // patch_install:     {"kb_ids": ["KB..."], "description": "..."}
    // manual_review:     {"reason": "...", "description": "..."}
    // unsupported:       {"reason": "..."}
  },
  "confidence": "high|medium|low"
}
判定规则：
- Windows 月度累积更新/KB 补丁 → patch_install，从 Results/Title 提取 KB 号
- 软件版本过低需升级 → software_upgrade
- EOL/停止支持的软件（如旧版 Office、旧版浏览器）→ manual_review（需业务确认）
- 纯注册表/配置加固（如 SMB signing）→ registry_fix（若涉及多个键，必须用 changes 数组逐键列出，详见上方 schema 说明）
- 无法明确判断 → manual_review，confidence 给 low"""


async def llm_parse_finding(db, qid: str, title: str, solution: str,
                            results: str = "") -> dict:
    """
    调用 openclaw 网关解析漏洞条目。
    返回 {"fix_type":..., "action":{...}, "confidence":...}
    失败时返回 manual_review + low（由上层升级为 high 风险）。
    """
    fallback = {
        "fix_type": "manual_review",
        "action": {"reason": "LLM 解析失败或不可用，需人工判断", "description": title[:200]},
        "confidence": "low",
    }
    from app.services.settings_service import get_ai_settings
    ai_cfg = await get_ai_settings(db)

    if not ai_cfg["llm_enabled"] or not ai_cfg["openclaw_token"]:
        fallback["action"]["reason"] = "LLM 未启用，需人工判断"
        return fallback

    user_prompt = (
        f"QID: {qid}\nTitle: {title}\n"
        f"Results(节选): {(results or '')[:800]}\n"
        f"Solution: {(solution or '')[:2000]}"
    )
    # 管理员在「连接设置」中配置的背景资料 / 人格设定，拼接到系统提示末尾
    system_content = LLM_SYSTEM_PROMPT
    if ai_cfg.get("openclaw_prompt"):
        system_content += (
            "\n\n# 企业背景与修复约束（管理员设定，必须严格遵守）\n"
            + ai_cfg["openclaw_prompt"]
        )
    try:
        async with httpx.AsyncClient(timeout=ai_cfg["openclaw_timeout"]) as client:
            resp = await client.post(
                f"{ai_cfg['openclaw_url'].rstrip('/')}/chat/completions",
                headers={"Authorization": f"Bearer {ai_cfg['openclaw_token']}"},
                json={
                    "model": ai_cfg["openclaw_model"],
                    "messages": [
                        {"role": "system", "content": system_content},
                        {"role": "user", "content": user_prompt},
                    ],
                    "temperature": 0,
                    "stream": False,
                },
            )
            resp.raise_for_status()
            content = resp.json()["choices"][0]["message"]["content"]
        parsed = _extract_json(content)
        if not parsed or parsed.get("fix_type") not in FIX_TYPES:
            logger.warning("LLM 输出无法解析为合法 JSON（QID=%s）：%.200s", qid, content)
            return fallback
        parsed.setdefault("action", {})
        if parsed.get("confidence") not in ("high", "medium", "low"):
            parsed["confidence"] = "low"
        return parsed
    except Exception as e:
        logger.warning("LLM 调用失败（QID=%s）：%s", qid, e)
        return fallback


def _extract_json(text: str) -> Optional[dict]:
    """从 LLM 输出中提取 JSON（容忍 ```json 代码块 / 前后杂文本）"""
    text = text.strip()
    m = re.search(r"```(?:json)?\s*(\{.*?\})\s*```", text, re.S)
    if m:
        text = m.group(1)
    else:
        # 找第一个 { 到最后一个 }
        s, e = text.find("{"), text.rfind("}")
        if s >= 0 and e > s:
            text = text[s:e + 1]
    try:
        obj = json.loads(text)
        return obj if isinstance(obj, dict) else None
    except (json.JSONDecodeError, ValueError):
        return None


# ── 主解析流程（后台任务）─────────────────────────────────────────────────────
def _norm(v) -> str:
    """单元格值 → 去空白字符串（并移除 x00 等 asyncpg 不可翻译字符）"""
    if v is None:
        return ""
    s = str(v).strip()
    # 移除 null byte 和其他 PostgreSQL/asyncpg 拒收的控制字符
    return s.replace('\x00', '').replace('\u0000', '')


# 规范字段 → 可接受的表头别名（已归一化：去空白 + 小写）
# 覆盖：英文原名、Qualys 标准名、中文常见写法
HEADER_ALIASES = {
    "ip":              ["ip", "ip地址", "主机ip", "资产ip", "ip address", "host ip", "host"],
    "dns_calculation": ["dns_calculation", "dns", "dnsname", "dns name", "dns hostname",
                        "主机名", "hostname", "计算机名", "computer name", "设备名", "设备"],
    "qid":             ["qid", "漏洞id", "漏洞编号", "漏洞序号", "vuln id", "qualys id", "qualys qid"],
    "title":           ["title", "漏洞标题", "标题", "漏洞名称", "漏洞名", "vulnerability",
                        "vuln title", "name", "漏洞"],
    "results":         ["results", "结果", "扫描结果", "扫描详情", "result", "details"],
    "solution":        ["solution", "解决方案", "修复建议", "修复方案", "建议", "处理建议",
                        "fix", "remediation", "处理方法", "处置建议"],
}


def _match_header(name) -> str | None:
    """归一化表头 → 规范字段名；匹配不到返回 None（支持别名）"""
    k = _norm(name).lower()
    if k in EXPECTED_HEADERS:
        return k
    for canon, variants in HEADER_ALIASES.items():
        if k in variants:
            return canon
    return None


def read_xlsx_rows(file_path: str) -> list[dict]:
    """
    读取 xlsx，自动定位表头行：
      - 跨所有工作表，在前 10 行内扫描匹配字段最多的表头行
        （容忍封面/标题/汇总行，以及首个工作表为空白的情况）
      - 按表头映射列（大小写/别名不敏感、容忍乱序）
    返回 [{ip, dns_name, qid, title, results, solution}, ...]，跳过 QID 为空的行。
    """
    wb = load_workbook(file_path, read_only=True, data_only=True)
    try:
        best = None  # (score, col_map, rows, header_idx)
        for ws in wb.worksheets:
            rows = list(ws.iter_rows(values_only=True))
            if not rows:
                continue
            for hidx in range(0, min(10, len(rows))):
                header = rows[hidx]
                col_map: dict[str, int] = {}
                for idx, name in enumerate(header or []):
                    field = _match_header(name) if name is not None else None
                    if field and field not in col_map.values():
                        col_map[field] = idx
                score = len(col_map)
                if best is None or score > best[0]:
                    best = (score, col_map, rows, hidx)

        if best is None or best[0] == 0:
            raise ValueError(
                "未识别到表头行（所有工作表前 10 行均无已知列）。"
                f"期望列：{', '.join(EXPECTED_HEADERS)}"
            )

        score, col_map, rows, hidx = best
        missing = [h for h in EXPECTED_HEADERS if h not in col_map]
        if missing:
            raise ValueError(
                f"缺少必需列：{', '.join(missing)}。"
                f"已识别列：{', '.join(col_map.keys())}（表头位于第 {hidx + 1} 行）。"
                f"期望完整列：{', '.join(EXPECTED_HEADERS)}"
            )

        out = []
        for raw in rows[hidx + 1:]:
            if raw is None:
                continue
            get = lambda k: _norm(raw[col_map[k]]) if col_map[k] < len(raw) else ""
            qid = get("qid")
            # 跳过全空行 / 无 QID 行
            if not qid:
                continue
            out.append({
                "ip":       get("ip"),
                "dns_name": get("dns_calculation"),
                "qid":      qid,
                "title":    get("title"),
                "results":  get("results"),
                "solution": get("solution"),
            })
        return out
    finally:
        wb.close()

async def parse_import(import_id: int, file_path: str):
    """
    后台解析任务入口（BackgroundTasks 调用，自带独立 Session）。
    任何异常都会把 import 标记为 failed 并记录 error_message。
    """
    async with AsyncSessionLocal() as db:
        imp = (await db.execute(
            select(VulnScanImport).where(VulnScanImport.id == import_id)
        )).scalar_one_or_none()
        if not imp:
            logger.error("parse_import: import %s 不存在", import_id)
            return
        try:
            imp.status = "parsing"
            await db.commit()

            rows = read_xlsx_rows(file_path)
            imp.row_count = len(rows)
            await db.commit()

            asset_index = await load_asset_index(db)
            # 批内 LLM 结果缓存：同一 QID 只解析一次
            llm_cache: dict[str, dict] = {}

            for row in rows:
                await _process_row(db, imp, row, asset_index, llm_cache)
                imp.processed_count += 1
                if imp.processed_count % 10 == 0:
                    await db.commit()   # 阶段性提交，前端可轮询进度

            imp.status = "completed"
            await db.commit()
            logger.info("import %s 解析完成：%s 行", import_id, imp.row_count)
        except Exception as e:
            logger.exception("import %s 解析失败", import_id)
            await db.rollback()
            imp = (await db.execute(
                select(VulnScanImport).where(VulnScanImport.id == import_id)
            )).scalar_one_or_none()
            if imp:
                imp.status = "failed"
                imp.error_message = str(e)[:2000]
                await db.commit()
        finally:
            # 注意：不再删除原始 xlsx（已持久化到 VULN_UPLOAD_DIR/stored/），
            # 以支持「重新解析」功能复用同一份文件。
            pass


async def _process_row(db: AsyncSession, imp: VulnScanImport, row: dict,
                       asset_index: dict, llm_cache: dict):
    """单行：落 finding → 匹配资产 → 依规则/LLM 生成 task"""
    asset_id, confidence = match_asset(asset_index, row["dns_name"], row["ip"])

    finding = VulnFinding(
        import_id=imp.id,
        ip=row["ip"] or None,
        dns_name=row["dns_name"] or None,
        qid=row["qid"],
        title=row["title"] or None,
        results_raw=row["results"] or None,
        solution_raw=row["solution"] or None,
        asset_id=asset_id,
        match_confidence=confidence,
    )
    db.add(finding)
    await db.flush()   # 拿 finding.id

    # 1) 查规则库
    rule = (await db.execute(
        select(RemediationRule).where(RemediationRule.qid == row["qid"])
    )).scalar_one_or_none()

    if rule and rule.status in ("active", "draft"):
        fix_type = rule.fix_type
        action = dict(rule.action_template or {})
        action.setdefault("description", (row["title"] or "")[:200])
        risk = rule.default_risk_level or classify_risk(fix_type, row["title"], row["solution"])
        source_note = f"rule:{rule.id}" + ("(draft)" if rule.status == "draft" else "")
    elif rule and rule.status == "disabled":
        # 规则被停用：不自动生成建议，转人工
        fix_type = "manual_review"
        action = {"reason": f"QID {row['qid']} 的规则已被停用，需人工确认处理方式",
                  "description": (row["title"] or "")[:200]}
        risk = "high"
        source_note = f"rule:{rule.id}(disabled)"
    else:
        # 2) LLM 兜底（批内同 QID 复用）
        if row["qid"] in llm_cache:
            parsed = llm_cache[row["qid"]]
        else:
            parsed = await llm_parse_finding(db, row["qid"], row["title"],
                                             row["solution"], row["results"])
            llm_cache[row["qid"]] = parsed
            # 写规则草稿（人工转正后生效；同 QID 只写一次）
            await _save_draft_rule(db, row["qid"], parsed, row["title"])
        fix_type = parsed["fix_type"]
        action = dict(parsed.get("action") or {})
        risk = classify_risk(fix_type, row["title"], row["solution"],
                             llm_confidence=parsed.get("confidence"))
        source_note = f"llm({parsed.get('confidence', '?')})"

    action["_source"] = source_note   # 溯源信息，前端可读化展示时显示

    task = RemediationTask(
        finding_id=finding.id,
        asset_id=asset_id,
        fix_type=fix_type,
        action_json=action,
        risk_level=risk,
        auto_approve=(risk == "low" and asset_id is not None),
        status="pending",
    )
    db.add(task)
    await db.flush()   # 拿 task.id

    # software_upgrade：生成任务时立即尝试匹配安装包（packages 库晚补充则可点「重新匹配」）
    if fix_type == "software_upgrade":
        from app.services.package_match import match_package
        sw = action.get("software") or action.get("download_hint") or ""
        tv = action.get("target_version") or ""
        pkg = await match_package(db, sw, tv)
        if pkg:
            task.matched_package_id = pkg.id
            await db.flush()


async def _save_draft_rule(db: AsyncSession, qid: str, parsed: dict, title: str):
    """LLM 解析结果 → status=draft 规则草稿（qid 已存在则跳过）"""
    exists = (await db.execute(
        select(RemediationRule.id).where(RemediationRule.qid == qid)
    )).scalar_one_or_none()
    if exists:
        return
    db.add(RemediationRule(
        qid=qid,
        fix_type=parsed["fix_type"],
        action_template=parsed.get("action") or {},
        default_risk_level=classify_risk(parsed["fix_type"], title,
                                         llm_confidence=parsed.get("confidence")),
        status="draft",
        source="llm",
        notes=f"LLM 自动生成（confidence={parsed.get('confidence', '?')}），标题：{title[:120]}",
    ))
